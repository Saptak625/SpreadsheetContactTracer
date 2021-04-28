import qrcode
import openpyxl
from openpyxl import load_workbook
import os
import smtplib, ssl
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import datetime
from PIL import Image, ImageFont, ImageDraw 
import imaplib
import email
from email.header import decode_header
import webbrowser
from dbFunctions import addNewContactTraceEntry
import datetime

#QR Code Generator
def generateQRCode(msg, filename):
  img = qrcode.make(msg)
  img.save(filename)
  img = Image.open(filename).convert('RGBA')
  draw = ImageDraw.Draw(img)
  # font = ImageFont.truetype(<font-file>, <font-size>)
  font = ImageFont.truetype("Roboto-Regular.ttf", 30)
  # draw.text((x, y),"Sample Text",(r,g,b))
  queryParameters = msg.split('?')[1].split('&')
  text = queryParameters[0].split('=')[1] + " - " + queryParameters[1].split('=')[1]
  draw.text((0, 0), text, (0,0,0), font=font)
  img.save(filename)

#Emailer
def createNewEmailMessage(EMAIL_ADDRESS, listOfContacts, subject, content, xlsxFileName):
  msg = EmailMessage()
  msg['Subject'] = subject
  msg['From'] = EMAIL_ADDRESS
  msg['To'] = ', '.join(listOfContacts)
  msg.set_content(content)
  
  with open(xlsxFileName, 'rb') as f:
    file_data = f.read()
    file_name = f.name

  msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename = file_name)
  return msg

def sendEmailWithXlsxAttachment(EMAIL_ADDRESS, PASSWORD, messages):
  context = ssl.create_default_context()
  try:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls(context=context)
    server.ehlo()
    server.login(EMAIL_ADDRESS, PASSWORD)
    for msg in messages:
      server.send_message(msg)
  except Exception as e:
    print(e)
  finally:
    server.quit()

def fetchOldExcelSheets():
  # account credentials
  username = 'dasdcontacttracer.noreply@gmail.com'
  password = 'iaqbetokntxxhxvi'

  # create an IMAP4 class with SSL 
  imap = imaplib.IMAP4_SSL("imap.gmail.com")
  # authenticate
  imap.login(username, password)

  status, messages = imap.select('"[Gmail]/Sent Mail"')

  # total number of emails
  messages = int(messages[0])
  # number of top emails to fetch
  N = messages

  for i in range(messages, messages-N, -1):
      # fetch the email message by ID
      res, msg = imap.fetch(str(i), "(RFC822)")
      for response in msg:
          if isinstance(response, tuple):
              # parse a bytes email into a message object
              msg = email.message_from_bytes(response[1])
              # decode the email subject
              subject, encoding = decode_header(msg["Subject"])[0]
              if 'Report for Room' not in subject:
                continue
              # if the email message is multipart
              if msg.is_multipart():
                  # iterate over email parts
                  for part in msg.walk():
                      # extract content type of email
                      content_type = part.get_content_type()
                      content_disposition = str(part.get("Content-Disposition"))
                      try:
                          # get the email body
                          body = part.get_payload(decode=True).decode()
                      except:
                          pass
                      if content_type == "text/plain" and "attachment" not in content_disposition:
                          # print text/plain emails and skip attachments
                          print(body)
                      elif "attachment" in content_disposition:
                          # download attachment
                          filename = part.get_filename()
                          if filename:
                              folder_name = 'Old Excels'
                              if not os.path.isdir(folder_name):
                                  # make a folder for this email (named after the subject)
                                  os.mkdir(folder_name)
                              filepath = os.path.join(folder_name, filename)
                              # download attachment and save it
                              open(filepath, "wb").write(part.get_payload(decode=True))
  # close the connection and logout
  imap.close()
  imap.logout()

def generateExcelSheet(results, name, numOfSeats):
    #Add Extra Empty Records
    valuesWithRecords = {i[3]: None for i in results}
    for desk in range(numOfSeats):
        if desk+1 not in valuesWithRecords:
            results.append(['', '', name, desk+1, ''])

    #Sort Results by desks
    def sortByDeskNumber(e):
      return e[3]
    results.sort(key=sortByDeskNumber)  

    #Create New Excel File
    workbook = Workbook()
    sheet = workbook.active

    #Make Header
    headerTitles = ['Desk', 'Name', 'Email', 'Time']
    ordOfA = ord('A')
    for i, header in enumerate(headerTitles):
      sheet[f'{chr(ordOfA+i)}1'] = header
      sheet[f'{ chr(ordOfA+i) }1'].font = Font(bold=True)

    #Make entries one by one
    deskNumbers = {}
    lastNumber = None
    for i, entry in enumerate(results):
        sheet[f'A{i+2}'] = entry[3]
        sheet[f'A{i+2}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'B{i+2}'] = entry[0]
        sheet[f'C{i+2}'] = entry[1]
        sheet[f'D{i+2}'] = entry[4]
        if lastNumber != entry[3]:
            if lastNumber != None:
                deskNumbers[lastNumber].append(i+2-1)
            lastNumber = entry[3]
            deskNumbers[lastNumber] = [i+2]
    if lastNumber != None:
        deskNumbers[lastNumber].append(len(results)-1+2)
    #Merge Desk Number
    for key in deskNumbers:
      value = deskNumbers[key]
      if value[0] != value[1]:
        #Merge cells
        sheet.merge_cells(f'A{value[0]}:A{value[1]}')
        sheet[f'A{value[0]}']

    #Save File once edits are made
    workbook.save(filename=f"{name}_{str(datetime.datetime.now()).split(' ')[0]}.xlsx")
    print('Excel Created')
    return f"{name}_{str(datetime.datetime.now()).split(' ')[0]}.xlsx"

def extractExcelRecords(filepath):
  #Get columns by (name TEXT, email TEXT, physicalclassroom TEXT, deskNumber INTEGER, entryTime TEXT)  
  # to open the workbook 
  # workbook object is created
  wb_obj = openpyxl.load_workbook(filepath)
  sheet_obj = wb_obj.active

  for i in range(sheet_obj.max_row-1):
    data = []
    for j in range(1, sheet_obj.max_column + 1):
      data.append(sheet_obj.cell(row = i+2, column = j).value)
    if None not in data:
      dateString = filepath.split('/')[1].split('_')[1].replace('.xlsx', '') + ' ' + data[3]
      data = [data[1], data[2], filepath.split('/')[1].split('_')[0], data[0], datetime.datetime(int(dateString.split(' ')[0].split('-')[0]), int(dateString.split(' ')[0].split('-')[1]), int(dateString.split(' ')[0].split('-')[2]), int(dateString.split(' ')[1].split(':')[0]), int(dateString.split(' ')[1].split(':')[1]), 0, 0)]
      addNewContactTraceEntry(data)
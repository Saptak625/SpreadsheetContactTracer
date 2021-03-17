import qrcode
from openpyxl import load_workbook

#Code to create new spreadsheet
# wb = Workbook()
# wb.save(filename="attendence.xlsx")

# # Data can be assigned directly to cells
# ws['A1'] = 42
#QR Code Generator
def generateQRCode(msg):
  img = qrcode.make(msg)
  img.save("QRCode.png")

#Excel Editor
def editSpreadsheet(cell, value):
  workbook = load_workbook(filename="attendence.xlsx")
  sheet = workbook.active
  sheet[cell] = value